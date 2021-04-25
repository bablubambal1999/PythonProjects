import pyautogui 
import time
import random
import webbrowser
import urllib.parse
import openpyxl
import keyboard

## get data from excel file function 
def getExcelData(fileName):
    workbook = openpyxl.load_workbook(fileName)
    active_sheet = workbook.active.title;
    sheet_object = workbook[active_sheet]
    rows = sheet_object.max_row
    columns = sheet_object.max_column
    arr_numbers = []
    arr_messages = []
    arr = []
    for i in range(1,rows+1):
        # for j in range(1,columns+1):
        #     #print(sh.cell(i,j).value,)
        arr_numbers.append(str(sheet_object.cell(i,1).value))
        arr_messages.append(sheet_object.cell(i,2).value)
    arr.append(arr_numbers)
    arr.append(arr_messages)
    return arr;
#time.sleep(5)

    

phoneNumbers,messages = getExcelData('data.xlsx')

print("Sending Message Starting")

for i in range(len(phoneNumbers)):
    message=urllib.parse.quote(messages[i])
    url = f"https://api.whatsapp.com/send/?phone=91{phoneNumbers[i]}&text={message}"
    webbrowser.open(url)
    time.sleep(3)


    print(url)
    keyboard.send('enter')
    time.sleep(1)
    # pyautogui.keyDown('enter')
    print("entre clicked")



    
    # pyautogui.keyDown('enter')
    # time.sleep(2)
    print("message send to ",i)
print("All messages Send")