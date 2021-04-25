# from openpyxl import Workbook
# wb = Workbook("data.xlsx")

# ws = wb.active
# ws1 = wb.create_sheet("Mysheet") 
# print("done")
# wb.save("excelsheet.xlsx")
import openpyxl
# wb = openpyxl.load_workbook('data.xlsx')
# she =wb.active.title;
# print(she)
# print(wb.sheetnames)
# print(wb['Sheet1']['A1'].value)
# print(wb['Sheet1']['B1'].value)
# sh  = wb['Sheet1']
# rows = sh.max_row
# columns = sh.max_column
# print(rows,columns)
# arr = []
# for i in range(1,rows+1):
#  for j in range(1,columns+1):
#      print(sh.cell(i,j).value,)
#      arr.append(sh.cell(i,j).value)
# # print("\n")
# print(arr)
def getExcelData(fileName):
    workbook = openpyxl.load_workbook(fileName)
    active_sheet = workbook.active.title;
    sheet_object = workbook[active_sheet]
    rows = sheet_object.max_row
    columns = sheet_object.max_column
    arr_numbers = []
    arr_messages = []
    arr = []
    for i in range(1,rows+1):
        # for j in range(1,columns+1):
        #     #print(sh.cell(i,j).value,)
        arr_numbers.append(str(sheet_object.cell(i,1).value))
        arr_messages.append(sheet_object.cell(i,2).value)
    arr.append(arr_numbers)
    arr.append(arr_messages)
    return arr;
print(getExcelData('data.xlsx'))
numbers,message = getExcelData('data.xlsx')
print(numbers, type(numbers[1]))
print(message,type(message[1]))
# import keyboard

# import time
# time.sleep(5)
# for i in range(10):
#     print("hello")
